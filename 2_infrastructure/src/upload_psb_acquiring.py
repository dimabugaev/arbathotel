import my_utility
import xlrd
import io
import requests
import json

from openpyxl import load_workbook

def read_xls_sheet(cursor, column_mapping: dict, key_column_name: str, sheet, field_index_misstring_sign: int, permanent_table_name:str, file_key):

    for row_index in range(1, sheet.nrows):

        try:
            #print(sheet.cell_value(row_index, field_index_misstring_sign))
            int(sheet.cell_value(row_index, field_index_misstring_sign))
        except:
            continue

        column_names = []
        column_values = []
        update_assignments = []
        for index, column_name in column_mapping.items():
            cell_value = sheet.cell_value(row_index, index)
            column_names.append(column_name)
            column_values.append(cell_value)
            update_assignments.append(f"{column_name} = %s")

        column_names.append('file_key')
        column_values.append(file_key)
        update_assignments.append(f"file_key = %s")

        insert_query = """
            INSERT INTO {table_name} ({column_list})
            VALUES ({value_placeholders})
            ON CONFLICT ({key_column})
            DO UPDATE SET {update_assignments}, date_update = current_timestamp;
        """.format(
            table_name=permanent_table_name,
            column_list=','.join(column_mapping.values()) + ',file_key',
            value_placeholders=','.join(['%s'] * (len(column_mapping) + 1)),
            key_column=key_column_name,
            update_assignments=','.join(update_assignments)
        )

        cursor.execute(insert_query, column_values + column_values)

def read_xlsx_sheet(cursor, column_mapping: dict, key_column_name: str, sheet, field_index_misstring_sign: int, permanent_table_name:str, file_key):

    for row in sheet.iter_rows(min_row=2, values_only=True):

        try:
            #print(row[field_index_misstring_sign])
            int(row[field_index_misstring_sign])
        except:
            continue

        column_names = []
        column_values = []
        update_assignments = []
        for index, column_name in column_mapping.items():
            cell_value = row[index]
            #if isinstance(cell_value, str):
            #    cell_value = cell_value.replace('_x000D_', '')

            column_names.append(column_name)
            column_values.append(cell_value)
            update_assignments.append(f"{column_name} = %s")

        column_names.append('file_key')
        column_values.append(file_key)
        update_assignments.append(f"file_key = %s")

        insert_query = """
            INSERT INTO {table_name} ({column_list})
            VALUES ({value_placeholders})
            ON CONFLICT ({key_column})
            DO UPDATE SET {update_assignments}, date_update = current_timestamp;
        """.format(
            table_name=permanent_table_name,
            column_list=','.join(column_mapping.values()) + ',file_key',
            value_placeholders=','.join(['%s'] * (len(column_mapping) + 1)),
            key_column=key_column_name,
            update_assignments=','.join(update_assignments)
        )

        cursor.execute(insert_query, column_values + column_values)

def do_normal_acquiring(cursor, sheet, file_key):
    
    column_mapping = {
        0: 'contract_name',
        1: 'device_name',
        2: 'device_number',
        3: 'device_addr',
        4: 'currency',
        5: 'payment_system',
        6: 'card_number',
        7: 'operation_data',
        8: 'processing_data',
        9: 'operation_sum',
        10: 'commission',
        11: 'to_transaction',
        12: 'rpn',
        13: 'operation_type',
        14: 'original_sum',
        15: 'original_currency'
    }

    if sheet.ncols > 16:
        column_mapping[16] = 'order_number'
        column_mapping[17] = 'description'   

    read_xls_sheet(cursor, column_mapping, 'rpn, operation_type', sheet, 12, 'banks_raw.psb_acquiring_term', file_key)

def do_qr_original(cursor, sheet, file_key):
    
    column_mapping = {
        0: 'date_time',
        1: 'id_payment',
        2: 'id_qr',
        3: 'payer_bank',
        4: 'payer_name',
        5: 'payer_account',
        6: 'recipient_inn',
        7: 'recipient_name',
        8: 'terminal_number',
        9: 'tsp_name',
        10: 'tsp_addr',
        11: 'recipient_account',
        12: 'mss',
        13: 'operation_sum',
        14: 'operation_com',
        15: 'to_tramsaction',
        16: 'currency',
        17: 'about_payment',
        18: 'description'
    }

    read_xlsx_sheet(cursor, column_mapping, 'id_payment', sheet, 13, 'banks_raw.psb_acquiring_qr', file_key)


def do_qr_refund(cursor, sheet, file_key):
    
    column_mapping = {
        0: 'date_time_original',
        1: 'date_time_refund',
        2: 'id_payment_refund',
        3: 'id_payment_original',
        4: 'terminal_number',
        5: 'tsp_addr',
        6: 'recipient_name',
        7: 'recipient_refund_account',
        8: 'payer_account',
        9: 'payer_tsp_name',
        10: 'tsp_id',
        11: 'refund_sum',
        12: 'original_sum',
        13: 'about_refund_payment',
        14: 'about_original_payment',
        15: 'payer_refund_bank',
        16: 'payer_original_bank',
        17: 'order_number'
    }

    read_xlsx_sheet(cursor, column_mapping, 'id_payment_refund', sheet, 12, 'banks_raw.psb_acquiring_qr_refund', file_key)

########################### HERE I STOPED
def get_payments(host, payment_id, login, password):

    session = requests.Session()
    url = f"https://{host}/info/options/byid/?id={payment_id}"

    session.headers.update({
        'Authorization': f'Basic {login}:{password}'
    })


    payment_data = {}

    with session.get(url) as response:
        payment_data = json.loads(response.text) 

    return payment_data

def upload_paykeeper_data(connect):
    update_plan_query = """
        with max_date_update as (
            select 
                max(date_update) last_writed
            from
                banks_raw.paykeeper_payments
        )
        ,internet_payments as (
            select
                contract_name,
                rpn,
                order_number,
                original_sum
            from
                banks_raw.psb_acquiring_term
            where 
                ((select last_writed from max_date_update) is null 
                or date_update > (select last_writed from max_date_update)) and order_number is not null
        ) 
        select distinct
            ip.payment_id,
            s.source_name host,
            s.source_username,
            s.source_password,
            s.source_id 
        from 
            (select 
                order_number payment_id,
                contract_name
            from
                internet_payments
            union all
            select
                pp.id,
                s.source_external_key 
            from
                banks_raw.paykeeper_payments pp join operate.sources s 
                on pp.orderid is null and pp.source_id = s.id 
	        ) ip left join operate.sources s on s.source_type = 6 and ip.contract_name = s.source_external_key;
    """

    cursor = connect.cursor()
    cursor.execute(update_plan_query)
    rows = cursor.fetchall()
    for row in rows:
        payment_info = get_payments(row[1], row[0], row[2], row[3])
        update_payment_in_rds(cursor, row[4], row[0], payment_info)
        connect.commit()
    cursor.close()
    return True



def update_payment_in_rds(cursor, source_id, payment_id, payment_info):
    payment_map = {
            'orderid':'orderid',
            'client_email':'client_email',
            'cart':'cart',
            'client_ip':'client_ip',
            'user_agent':'user_agent',
            'fop_receipt_key':'fop_receipt_key',
            'CARD_NUMBER':'card_number',
            'EXP':'exp',
            'RRN':'rpn',
            'INT_REF':'int_ref',
            'NAME':'name',
            'APPROVAL_CODE':'approval_code',
            'RC':'rc',
            'RCTEXT':'rctext',
            'MESSAGE':'message',
            'ACTION':'action',
            'fop_uuid':'fop_uuid',
            'fop_tax_comment':'fop_tax_comment',
            'fop_fpd':'fop_fpd',
            'fop_receipt_number':'fop_receipt_number',
            'fop_url':'fop_url',
            'fop_fn':'fop_fn',
            'fop_fnd':'fop_fnd',
            'fop_rnkkt':'fop_rnkkt',
            'fop_shift_number':'fop_shift_number' 
        }

    column_names = 'source_id,id'
    column_values_params = '%s,%s'
    column_values = [source_id, payment_id]
    update_assignments = 'date_update = current_timestamp,source_id = EXCLUDED.source_id'
    for k, v in payment_map:
        value = payment_info.get(k)
        if value:
            column_values.append(value)
            column_names += f',{v}'
            column_values_params += ',%s'
            update_assignments += f',{v} = EXCLUDED.{v}' 

    insert_query = f"""
            INSERT INTO banks_raw.paykeeper_payments ({column_names})
            VALUES ({column_values_params})
            ON CONFLICT (id)
            DO UPDATE SET {update_assignments};
        """
    cursor.execute(insert_query, tuple(column_values))
    return True

def upload_data_to_rds():
    source_data = my_utility.get_email_and_storage_data()
    s3client = source_data['s3client']
    source_bucket = source_data['s3_bucket_for_attachments']
    source_prefix = 'dev/psb-acquiring/income/'
    destination_prefix = 'dev/psb-acquiring/done/'

    s3_objects = s3client.list_objects_v2(Bucket=source_bucket, Prefix=source_prefix).get('Contents')
    conn = my_utility.get_db_connection()

    if s3_objects is not None:

        cursor = conn.cursor()

        for s3_object in s3_objects:
            s3_response = s3client.get_object(Bucket=source_bucket, Key=s3_object['Key'])
            print(s3_object['Key'])
            xls_data = s3_response['Body'].read()

            if s3_object['Key'].lower().endswith('_e.xls'):
                workbook = xlrd.open_workbook(file_contents=xls_data)
                sheet = workbook.sheet_by_index(0)
                do_normal_acquiring(cursor, sheet, s3_object['Key'])   
            elif s3_object['Key'].lower().endswith('.xlsx'):
                
                workbook = load_workbook(filename=io.BytesIO(xls_data))
                if len(workbook.sheetnames) > 1:
                    do_qr_original(cursor, workbook.get_sheet_by_name(workbook.sheetnames[0]), s3_object['Key'])
                    do_qr_refund(cursor, workbook.get_sheet_by_name(workbook.sheetnames[1]), s3_object['Key'])
            else:
                print(s3_object['Key'] + ' - bad format')
                continue


            conn.commit()
            destination_key = destination_prefix + s3_object['Key'].split('/')[-1]
            s3client.copy_object(Bucket=source_bucket, Key=destination_key, CopySource={'Bucket': source_bucket, 'Key': s3_object['Key']})
            
            s3client.delete_object(Bucket=source_bucket, Key=s3_object['Key'])

        print('Данные успешно обработаны и записаны в постоянную таблицу, файлы перемещены.')

        # Закрытие соединений
        cursor.close()
    else:
        print('New data is empty')

    upload_paykeeper_data(conn)
    conn.close()

def lambda_handler(event, context):
    upload_data_to_rds()

